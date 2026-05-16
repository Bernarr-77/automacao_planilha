from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from typing import List
from datetime import datetime 
from database import get_all
from io import BytesIO

def gerar_relatorio(dados: List ):
    wb = Workbook()

    planilha = wb.active

    planilha.append(['Cliente', 'Serviço', 'Data/Hora', 'Valor', 'Status'])

    fundo_verde = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    texto_branco = Font(color="FFFFFF", bold=True)

    for celula in planilha[1]:
        celula.fill = fundo_verde
        celula.font = texto_branco

    for _, cliente, servico, data_texto, valor, status in dados:
        data_convertida = datetime.fromisoformat(data_texto)
        planilha.append([cliente, servico, data_convertida, valor, status])
    planilha.column_dimensions["A"].width = 30  # Cliente
    planilha.column_dimensions["B"].width = 25  # Serviço
    planilha.column_dimensions["C"].width = 25  # Data/Hora
    planilha.column_dimensions["D"].width = 15  # Valor
    planilha.column_dimensions["E"].width = 15  # Status

    for linha in range(2, planilha.max_row + 1):    
        celula_data = planilha.cell(row=linha, column=3)
        celula_data.number_format = 'dd/mm/yyyy hh:mm:ss'
        celula_valor = planilha.cell(row=linha, column=4)
        celula_valor.number_format = '"R$ "#,##0.00'

    planilha.auto_filter.ref = planilha.dimensions
    fluxo = BytesIO()
    wb.save(fluxo)
    fluxo.seek(0)
    return fluxo

if __name__ == "__main__":
    todos_dados = get_all()
    gerar_relatorio(todos_dados)